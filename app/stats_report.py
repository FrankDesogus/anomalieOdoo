from __future__ import annotations

from datetime import datetime, time
from pathlib import Path
from typing import List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .stats_builder import ArrivalRecord, BucketCount, MonthlyStats, PeriodStats


# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------

_TITLE_FONT = Font(bold=True, size=13)
_LABEL_FONT = Font(bold=True)
_HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_VALID_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
_EXCLUDED_FILL = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")


def _auto_width(ws, max_width: int = 50) -> None:
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max(
            (len(str(cell.value)) for cell in col if cell.value is not None),
            default=0,
        )
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


def _fmt_time(t: Optional[time]) -> str:
    if t is None:
        return "N/D"
    return f"{t.hour:02d}:{t.minute:02d}"


def _fmt_pct(v: Optional[float]) -> str:
    if v is None:
        return "N/D"
    return f"{v:.1f}%"


def _apply_header_style(ws, row: int, n_cols: int) -> None:
    for c in range(1, n_cols + 1):
        cell = ws.cell(row, c)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


# ---------------------------------------------------------------------------
# Sheet: Sintesi
# ---------------------------------------------------------------------------

def _write_sintesi(wb: Workbook, stats: PeriodStats, employee_pool: List[Tuple[int, str]]) -> None:
    ws = wb.create_sheet("Sintesi")
    r = 1

    ws.cell(r, 1).value = "STATISTICHE ORARI DI INGRESSO"
    ws.cell(r, 1).font = _TITLE_FONT
    r += 2

    def kv(label: str, value) -> None:
        nonlocal r
        ws.cell(r, 1).value = label
        ws.cell(r, 1).font = _LABEL_FONT
        ws.cell(r, 2).value = value
        r += 1

    kv("Periodo", f"{stats.from_day} → {stats.to_day}")
    kv("Finestra oraria", f"{_fmt_time(stats.arrival_start)} – {_fmt_time(stats.arrival_end)}")
    kv("Ampiezza fasce", f"{stats.bucket_minutes} min")
    kv("Dipendenti nel pool", stats.employee_count)
    kv("Ingressi validi", stats.valid_count)
    kv("Ingressi fuori fascia (esclusi)", stats.excluded_count)
    kv("Media", _fmt_time(stats.mean_dt))
    kv("Mediana", _fmt_time(stats.median_dt))
    kv("Deviazione standard", f"{stats.std_minutes:.1f} min" if stats.std_minutes is not None else "N/D")
    kv("Fascia più frequente", stats.most_frequent_bucket or "N/D")
    kv("Ingressi entro 08:15", _fmt_pct(stats.pct_by_0815))
    kv("Ingressi entro 08:30", _fmt_pct(stats.pct_by_0830))
    kv("Ingressi entro 08:45", _fmt_pct(stats.pct_by_0845))
    kv("Report generato il", datetime.now().strftime("%Y-%m-%d %H:%M"))

    r += 2

    # Bucket table (used as chart data source)
    chart_start = r
    ws.cell(r, 1).value = "Fascia oraria"
    ws.cell(r, 2).value = "Ingressi validi"
    ws.cell(r, 1).font = _LABEL_FONT
    ws.cell(r, 2).font = _LABEL_FONT
    r += 1

    for b in stats.buckets:
        ws.cell(r, 1).value = b.label
        ws.cell(r, 2).value = b.count
        r += 1

    chart_end = r - 1

    if stats.valid_count > 0 and len(stats.buckets) > 0:
        chart = BarChart()
        chart.type = "col"
        chart.grouping = "clustered"
        chart.title = "Distribuzione ingressi validi per fascia"
        chart.y_axis.title = "N. ingressi"
        chart.x_axis.title = "Fascia"
        chart.width = 22
        chart.height = 13

        data = Reference(ws, min_col=2, min_row=chart_start, max_row=chart_end)
        cats = Reference(ws, min_col=1, min_row=chart_start + 1, max_row=chart_end)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, f"D{chart_start}")

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 30


# ---------------------------------------------------------------------------
# Sheet: Andamento mensile
# ---------------------------------------------------------------------------

def _write_andamento_mensile(
    wb: Workbook, monthly: List[MonthlyStats], arrival_start: time
) -> None:
    ws = wb.create_sheet("Andamento mensile")

    COLS = [
        "Mese", "Ingressi validi", "Esclusi fuori fascia",
        "Media", "Mediana", "Dev. std. (min)",
        "Fascia piu' frequente",
        "Entro 08:15", "Entro 08:30", "Entro 08:45",
    ]
    HELPER_COL = len(COLS) + 1  # minutes from arrival_start, used by chart

    ref_mins = arrival_start.hour * 60 + arrival_start.minute
    ref_label = arrival_start.strftime("%H:%M")

    # Header
    for c, h in enumerate(COLS, 1):
        ws.cell(1, c).value = h
    _apply_header_style(ws, 1, len(COLS))
    ws.cell(1, HELPER_COL).value = f"Min da {ref_label}"

    data_start = 2
    for i, m in enumerate(monthly):
        row = data_start + i
        ws.cell(row, 1).value = m.year_month
        ws.cell(row, 2).value = m.valid_count
        ws.cell(row, 3).value = m.excluded_count
        ws.cell(row, 4).value = _fmt_time(m.mean_dt)
        ws.cell(row, 5).value = _fmt_time(m.median_dt)
        ws.cell(row, 6).value = m.std_minutes if m.std_minutes is not None else ""
        ws.cell(row, 7).value = m.most_frequent_bucket or ""
        ws.cell(row, 8).value = _fmt_pct(m.pct_by_0815)
        ws.cell(row, 9).value = _fmt_pct(m.pct_by_0830)
        ws.cell(row, 10).value = _fmt_pct(m.pct_by_0845)
        # Minutes elapsed since arrival_start (e.g. 08:15 → 15 min, 08:30 → 30 min)
        if m.mean_dt is not None:
            mean_mins = m.mean_dt.hour * 60 + m.mean_dt.minute
            ws.cell(row, HELPER_COL).value = mean_mins - ref_mins

    data_end = data_start + len(monthly) - 1

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    if any(m.mean_dt is not None for m in monthly) and len(monthly) > 0:
        chart = LineChart()
        chart.title = f"Media mensile ingressi (min trascorsi dalle {ref_label})"
        chart.y_axis.title = f"Min da {ref_label}"
        chart.x_axis.title = "Mese"
        chart.width = 26
        chart.height = 13

        data = Reference(ws, min_col=HELPER_COL, min_row=1, max_row=data_end)
        cats = Reference(ws, min_col=1, min_row=data_start, max_row=data_end)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, f"A{data_end + 3}")

    _auto_width(ws)
    # Hide helper column — raw numbers, only needed for the chart
    ws.column_dimensions[get_column_letter(HELPER_COL)].hidden = True


# ---------------------------------------------------------------------------
# Sheet: Distribuzione mensile
# ---------------------------------------------------------------------------

def _write_distribuzione_mensile(wb: Workbook, monthly: List[MonthlyStats]) -> None:
    ws = wb.create_sheet("Distribuzione mensile")

    if not monthly:
        ws.cell(1, 1).value = "Nessun dato disponibile"
        return

    bucket_labels = [b.label for b in monthly[0].buckets]
    headers = ["Mese"] + bucket_labels
    for c, h in enumerate(headers, 1):
        ws.cell(1, c).value = h
    _apply_header_style(ws, 1, len(headers))

    data_start = 2
    for i, m in enumerate(monthly):
        row = data_start + i
        ws.cell(row, 1).value = m.year_month
        total = sum(b.count for b in m.buckets)
        for j, b in enumerate(m.buckets, 2):
            pct = round(100.0 * b.count / total, 1) if total > 0 else 0.0
            ws.cell(row, j).value = pct

    data_end = data_start + len(monthly) - 1

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions

    if len(bucket_labels) > 0 and any(m.valid_count > 0 for m in monthly):
        chart = BarChart()
        chart.type = "col"
        chart.grouping = "percentStacked"
        chart.title = "Distribuzione mensile per fascia (%)"
        chart.y_axis.title = "%"
        chart.x_axis.title = "Mese"
        chart.width = 28
        chart.height = 14

        n_buckets = len(bucket_labels)
        data = Reference(ws, min_col=2, max_col=1 + n_buckets, min_row=1, max_row=data_end)
        cats = Reference(ws, min_col=1, min_row=data_start, max_row=data_end)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, f"A{data_end + 3}")

    _auto_width(ws)


# ---------------------------------------------------------------------------
# Sheet: Dettaglio dati
# ---------------------------------------------------------------------------

def _write_dettaglio(wb: Workbook, records: List[ArrivalRecord]) -> None:
    ws = wb.create_sheet("Dettaglio dati")

    headers = [
        "Data", "Employee ID", "Dipendente",
        "Prima timbratura reale", "Incluso nella statistica", "Motivo", "Mese",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(1, c).value = h
    _apply_header_style(ws, 1, len(headers))

    for i, rec in enumerate(records):
        row = 2 + i
        ws.cell(row, 1).value = rec.day.isoformat()
        ws.cell(row, 2).value = rec.employee_id
        ws.cell(row, 3).value = rec.employee_name
        ws.cell(row, 4).value = rec.first_checkin.strftime("%H:%M:%S")
        ws.cell(row, 5).value = "Sì" if rec.included else "No"
        ws.cell(row, 6).value = rec.exclusion_reason
        ws.cell(row, 7).value = f"{rec.day.year:04d}-{rec.day.month:02d}"

        fill = _VALID_FILL if rec.included else _EXCLUDED_FILL
        for c in range(1, len(headers) + 1):
            ws.cell(row, c).fill = fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    _auto_width(ws)


# ---------------------------------------------------------------------------
# Sheet: Pool dipendenti
# ---------------------------------------------------------------------------

def _write_pool(wb: Workbook, employee_pool: List[Tuple[int, str]]) -> None:
    ws = wb.create_sheet("Pool dipendenti")

    for c, h in enumerate(["Employee ID", "Dipendente"], 1):
        ws.cell(1, c).value = h
    _apply_header_style(ws, 1, 2)

    for i, (emp_id, name) in enumerate(sorted(employee_pool, key=lambda x: (x[1] or "").lower())):
        ws.cell(2 + i, 1).value = emp_id
        ws.cell(2 + i, 2).value = name

    _auto_width(ws)


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def export_stats_excel(
    stats: PeriodStats,
    monthly: List[MonthlyStats],
    records: List[ArrivalRecord],
    employee_pool: List[Tuple[int, str]],
    output_path: Path,
) -> Path:
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    _write_sintesi(wb, stats, employee_pool)
    _write_andamento_mensile(wb, monthly, stats.arrival_start)
    _write_distribuzione_mensile(wb, monthly)
    _write_dettaglio(wb, records)
    _write_pool(wb, employee_pool)

    wb.save(output)
    return output
