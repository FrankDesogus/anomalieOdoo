"""
Test per app/report_paths.py e comportamenti correlati (stats_report, PermissionError).
"""
from __future__ import annotations

from datetime import date, time
from pathlib import Path
from unittest.mock import patch

import pytest

from app.report_paths import (
    build_default_report_filename,
    next_available_path,
    sanitize_filename_component,
)


# ---------------------------------------------------------------------------
# 1. build_default_report_filename — pool "Meccanica"
# ---------------------------------------------------------------------------

def test_filename_pool_meccanica():
    result = build_default_report_filename("Meccanica", date(2026, 1, 1), date(2026, 6, 9))
    assert result == "statistiche_ingressi_meccanica_2026-01-01_2026-06-09.xlsx"


# ---------------------------------------------------------------------------
# 2. build_default_report_filename — pool "Ufficio Tecnico"
# ---------------------------------------------------------------------------

def test_filename_pool_ufficio_tecnico():
    result = build_default_report_filename("Ufficio Tecnico", date(2026, 1, 1), date(2026, 6, 9))
    assert result == "statistiche_ingressi_ufficio_tecnico_2026-01-01_2026-06-09.xlsx"


# ---------------------------------------------------------------------------
# 3. Fallback selezione_manuale quando pool_name è None
# ---------------------------------------------------------------------------

def test_filename_fallback_none():
    result = build_default_report_filename(None, date(2026, 1, 1), date(2026, 6, 9))
    assert result == "statistiche_ingressi_selezione_manuale_2026-01-01_2026-06-09.xlsx"


def test_filename_fallback_empty_string():
    result = build_default_report_filename("", date(2026, 1, 1), date(2026, 6, 9))
    assert result == "statistiche_ingressi_selezione_manuale_2026-01-01_2026-06-09.xlsx"


# ---------------------------------------------------------------------------
# 4. sanitize — normalizzazione spazi
# ---------------------------------------------------------------------------

def test_sanitize_spaces():
    assert sanitize_filename_component("Ufficio Tecnico") == "ufficio_tecnico"


def test_sanitize_multiple_spaces():
    assert sanitize_filename_component("Ufficio  Tecnico") == "ufficio_tecnico"


# ---------------------------------------------------------------------------
# 5. sanitize — normalizzazione accenti
# ---------------------------------------------------------------------------

def test_sanitize_accents_a():
    assert sanitize_filename_component("Qualità") == "qualita"


def test_sanitize_accents_mixed():
    assert sanitize_filename_component("Qualità / Collaudo") == "qualita_collaudo"


# ---------------------------------------------------------------------------
# 6. sanitize — rimozione caratteri vietati Windows
# ---------------------------------------------------------------------------

def test_sanitize_colon():
    assert sanitize_filename_component("Pool: Amministrazione") == "pool_amministrazione"


def test_sanitize_forbidden_chars():
    result = sanitize_filename_component('Test<>:"/\\|?*')
    assert "<" not in result
    assert ">" not in result
    assert ":" not in result
    assert '"' not in result
    assert "/" not in result
    assert "\\" not in result
    assert "|" not in result
    assert "?" not in result
    assert "*" not in result


# ---------------------------------------------------------------------------
# 7. next_available_path — nome base se il file non esiste
# ---------------------------------------------------------------------------

def test_next_available_path_no_existing(tmp_path):
    p = tmp_path / "report.xlsx"
    assert next_available_path(p) == p


# ---------------------------------------------------------------------------
# 8. next_available_path — suffisso _02 quando il file base esiste
# ---------------------------------------------------------------------------

def test_next_available_path_02(tmp_path):
    p = tmp_path / "report.xlsx"
    p.touch()
    result = next_available_path(p)
    assert result == tmp_path / "report_02.xlsx"


# ---------------------------------------------------------------------------
# 9. next_available_path — suffisso _03 quando esistono base e _02
# ---------------------------------------------------------------------------

def test_next_available_path_03(tmp_path):
    base = tmp_path / "report.xlsx"
    v02 = tmp_path / "report_02.xlsx"
    base.touch()
    v02.touch()
    result = next_available_path(base)
    assert result == tmp_path / "report_03.xlsx"


# ---------------------------------------------------------------------------
# 10. next_available_path — non elimina i file precedenti
# ---------------------------------------------------------------------------

def test_next_available_path_preserves_existing(tmp_path):
    base = tmp_path / "report.xlsx"
    base.touch()
    result = next_available_path(base)
    # Il file originale esiste ancora
    assert base.exists()
    # Il nuovo path non esiste ancora
    assert not result.exists()


# ---------------------------------------------------------------------------
# 11. build_default_report_filename è deterministica (base per "ripristino")
# ---------------------------------------------------------------------------

def test_filename_is_deterministic():
    """
    Chiamare build_default_report_filename due volte con gli stessi parametri
    restituisce sempre lo stesso nome — garanzia del pulsante "Ripristina".
    """
    args = ("Meccanica", date(2026, 1, 1), date(2026, 6, 9))
    assert build_default_report_filename(*args) == build_default_report_filename(*args)


# ---------------------------------------------------------------------------
# 12. export_stats_excel propaga PermissionError senza nasconderla
# ---------------------------------------------------------------------------

def test_export_permission_error_propagates(tmp_path):
    from app.stats_builder import PeriodStats
    from app.stats_report import export_stats_excel
    from openpyxl import Workbook

    stats = PeriodStats(
        from_day=date(2026, 1, 1),
        to_day=date(2026, 1, 31),
        arrival_start=time(8, 0),
        arrival_end=time(9, 0),
        bucket_minutes=10,
        employee_count=1,
        valid_count=0,
        excluded_count=0,
        mean_dt=None,
        median_dt=None,
        std_minutes=None,
        most_frequent_bucket=None,
        pct_by_0815=None,
        pct_by_0830=None,
        pct_by_0845=None,
        buckets=[],
    )
    out = tmp_path / "test.xlsx"

    with patch.object(Workbook, "save", side_effect=PermissionError("file bloccato")):
        with pytest.raises(PermissionError):
            export_stats_excel(stats, [], [], [(1, "Test User")], out)


# ---------------------------------------------------------------------------
# 13. Il foglio Sintesi contiene la riga "Report generato il"
# ---------------------------------------------------------------------------

def test_sintesi_has_generated_at_row(tmp_path):
    from app.stats_builder import PeriodStats
    from app.stats_report import export_stats_excel
    from openpyxl import load_workbook as _lw

    stats = PeriodStats(
        from_day=date(2026, 1, 1),
        to_day=date(2026, 1, 31),
        arrival_start=time(8, 0),
        arrival_end=time(9, 0),
        bucket_minutes=10,
        employee_count=1,
        valid_count=0,
        excluded_count=0,
        mean_dt=None,
        median_dt=None,
        std_minutes=None,
        most_frequent_bucket=None,
        pct_by_0815=None,
        pct_by_0830=None,
        pct_by_0845=None,
        buckets=[],
    )
    out = tmp_path / "test.xlsx"
    export_stats_excel(stats, [], [], [(1, "Test User")], out)

    wb = _lw(out)
    ws = wb["Sintesi"]
    col_a_values = [
        str(ws.cell(r, 1).value)
        for r in range(1, ws.max_row + 1)
        if ws.cell(r, 1).value is not None
    ]
    assert any("Report generato il" in v for v in col_a_values), (
        f"'Report generato il' non trovato nel foglio Sintesi. Righe: {col_a_values}"
    )
