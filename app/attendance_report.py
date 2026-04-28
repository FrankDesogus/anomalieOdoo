from __future__ import annotations

from collections import defaultdict
from datetime import date
from pathlib import Path
from typing import Dict, Iterable, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import AttendanceSession


ROW_COLUMNS = [
    "Data",
    "ID Dipendente",
    "Nome Dipendente",
    "Stato",
    "Primo ingresso",
    "Ultima uscita",
    "Numero sessioni",
    "Note",
]


PRESENT_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
ABSENT_FILL = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")


def format_date(value):
    if not value:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")
    return str(value)


def format_time(value):
    if not value:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%H:%M:%S")
    return str(value)


def build_attendance_rows(
    day: date,
    employees: Iterable[Tuple[int, str]],
    sessions: Iterable[AttendanceSession],
) -> List[Dict[str, object]]:
    sessions_by_employee: Dict[int, List[AttendanceSession]] = defaultdict(list)
    for session in sessions:
        sessions_by_employee[session.employee_id].append(session)

    rows: List[Dict[str, object]] = []
    day_str = day.isoformat()

    sorted_employees = sorted(employees, key=lambda e: (e[1] or "").lower())
    for employee_id, employee_name in sorted_employees:
        emp_sessions = sorted(sessions_by_employee.get(employee_id, []), key=lambda s: s.check_in)

        if not emp_sessions:
            rows.append(
                {
                    "Data": day_str,
                    "ID Dipendente": employee_id,
                    "Nome Dipendente": employee_name,
                    "Stato": "Assente",
                    "Primo ingresso": None,
                    "Ultima uscita": None,
                    "Numero sessioni": 0,
                    "Note": "Nessuna timbratura registrata",
                }
            )
            continue

        has_open = any(s.check_out is None for s in emp_sessions)
        first_in = min(s.check_in for s in emp_sessions)
        closed_sessions = [s for s in emp_sessions if s.check_out is not None]
        last_out = max((s.check_out for s in closed_sessions), default=None)
        sessions_count = len(emp_sessions)

        if sessions_count > 1 and has_open:
            note = "Presenza con più sessioni, una ancora aperta"
        elif sessions_count > 1:
            note = "Presenza con più sessioni"
        elif has_open:
            note = "Presenza aperta"
        else:
            note = "Presenza già chiusa"

        rows.append(
            {
                "Data": day_str,
                "ID Dipendente": employee_id,
                "Nome Dipendente": employee_name,
                "Stato": "Presente",
                "Primo ingresso": first_in,
                "Ultima uscita": last_out,
                "Numero sessioni": sessions_count,
                "Note": note,
            }
        )

    return rows


def export_attendance_excel(day: date, rows: List[Dict[str, object]], output_path: Path | str) -> Path:
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = f"Presenze {day.isoformat()}"

    ws.append(ROW_COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in rows:
        excel_row = [
            format_date(row.get("Data")),
            row.get("ID Dipendente", ""),
            row.get("Nome Dipendente", ""),
            row.get("Stato", ""),
            format_time(row.get("Primo ingresso")),
            format_time(row.get("Ultima uscita")),
            row.get("Numero sessioni", ""),
            row.get("Note", ""),
        ]
        ws.append(excel_row)

        current_row = ws.max_row
        status = row.get("Stato")
        fill = PRESENT_FILL if status == "Presente" else ABSENT_FILL
        for cell in ws[current_row]:
            cell.fill = fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for idx, column_name in enumerate(ROW_COLUMNS, start=1):
        max_len = len(column_name)
        for row_idx in range(2, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=idx).value
            text = "" if value is None else str(value)
            max_len = max(max_len, len(text))
        ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 60)

    wb.save(output)
    return output
