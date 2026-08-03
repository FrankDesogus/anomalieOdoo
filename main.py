import argparse
import yaml
from datetime import date, datetime, time, timedelta
from pathlib import Path
from zoneinfo import ZoneInfo

from app.detector import detect_for_day
from app.fake_data import build_fake_sessions
from app.dedup import DedupStore
from app.mailer import send_mail
from app.odoo_client import OdooClient
from app.attendance_report import build_attendance_rows, export_attendance_excel
from app.stats_service import (
    ArrivalStatsParams,
    default_output_path,
    parse_date as _svc_parse_date,
    parse_time as _svc_parse_time,
    resolve_pool_from_ids,
    run as _svc_run,
    validate_params as _svc_validate_params,
)


def day_range_utc(day: date, tz_name: str):
    tz = ZoneInfo(tz_name)
    start_local = datetime.combine(day, time(0, 0)).replace(tzinfo=tz)
    end_local = start_local + timedelta(days=1)
    start_utc = start_local.astimezone(ZoneInfo("UTC")).replace(tzinfo=None)
    end_utc = end_local.astimezone(ZoneInfo("UTC")).replace(tzinfo=None)
    return start_utc.strftime("%Y-%m-%d %H:%M:%S"), end_utc.strftime("%Y-%m-%d %H:%M:%S")


def format_anomalies(anomalies):
    lines = []

    def friendly_line(a):
        day_s = str(a.day)

        if a.code == "LATE_CHECKIN":
            # usa evidence se c'è, altrimenti fallback sul testo già pronto
            t = None
            if isinstance(a.evidence, dict) and a.evidence.get("first_check_in"):
                # "2026-01-20T09:02:00+01:00" -> prendo HH:MM
                t = a.evidence["first_check_in"][11:16]
            if t:
                msg = f"Timbratura alle {t} (oltre le 09:00). Probabile ritardo o dimenticanza."
            else:
                msg = a.message + " Probabile ritardo o dimenticanza."

        elif a.code == "NO_CHECKIN_BY_0900":
            msg = "Nessuna timbratura di ingresso entro le 09:00. Probabile assenza o dimenticanza."

        elif a.code == "LUNCH_LATE_RETURN":
            msg = a.message + " Probabile rientro tardivo o mancata timbratura."

        elif a.code == "LUNCH_NO_RETURN":
            msg = a.message + " Probabile mancata timbratura di rientro."

        elif a.code == "AUTO_CLOSE_LATE":
            msg = a.message  # già user-friendly

        elif a.code == "NO_ATTENDANCE_ALL_DAY":
            msg = a.message  # già user-friendly

        else:
            # fallback generico
            msg = a.message

        return f"{a.employee_name}  {day_s}  {msg}"

    for a in anomalies:
        lines.append(friendly_line(a))

    return "\n".join(lines)


def month_days(yyyy_mm: str) -> list[date]:
    y_s, m_s = yyyy_mm.split("-")
    y, m = int(y_s), int(m_s)
    start = date(y, m, 1)
    if m == 12:
        end = date(y + 1, 1, 1) - timedelta(days=1)
    else:
        end = date(y, m + 1, 1) - timedelta(days=1)
    return [start + timedelta(days=i) for i in range((end - start).days + 1)]


def build_days_to_check(args) -> list[date]:
    today = date.today()

    if args.mode == "ops":
        # oggi + lookback (default 1 = ieri)
        n = max(0, int(args.lookback_days))
        return [today - timedelta(days=i) for i in range(n + 1)]

    if args.mode == "day":
        if not args.day:
            raise RuntimeError("mode=day richiede --day YYYY-MM-DD")
        d = date.fromisoformat(args.day)
        return [d - timedelta(days=1), d] if args.also_prev else [d]

    if args.mode == "month":
        if not args.month:
            raise RuntimeError("mode=month richiede --month YYYY-MM")
        return month_days(args.month)

    raise RuntimeError("mode non valido")


def resolve_target_day(day_arg: str | None, tz_name: str) -> date:
    if day_arg:
        return date.fromisoformat(day_arg)
    return datetime.now(ZoneInfo(tz_name)).date()


# ---------------------------------------------------------------------------
# arrival-stats helpers
# ---------------------------------------------------------------------------

def validate_arrival_args(args) -> tuple:
    """Return (from_day, to_day, arrival_start, arrival_end, bucket_minutes)."""
    today = date.today()
    from_day = _svc_parse_date(args.from_day) if args.from_day else date(today.year, 1, 1)
    to_day = _svc_parse_date(args.to_day) if args.to_day else today
    arrival_start = _svc_parse_time(args.arrival_start)
    arrival_end = _svc_parse_time(args.arrival_end)
    bm = args.bucket_minutes
    _svc_validate_params(from_day, to_day, arrival_start, arrival_end, bm)
    return from_day, to_day, arrival_start, arrival_end, bm


def _load_pool_from_excel_file(path: str) -> list:
    """Read employee IDs from an Excel pool file (columns: employee_id, include)."""
    from openpyxl import load_workbook as _lw
    wb = _lw(path, read_only=True)
    ws = wb.active
    headers = [
        str(c.value).strip().lower() if c.value is not None else ""
        for c in next(ws.iter_rows(min_row=1, max_row=1))
    ]
    try:
        id_col = headers.index("employee_id")
        include_col = headers.index("include")
    except ValueError as exc:
        raise ValueError(
            f"Il file pool '{path}' deve avere le colonne 'employee_id' e 'include': {exc}"
        )
    ids = []
    for row in ws.iter_rows(min_row=2):
        id_val = row[id_col].value
        inc_val = row[include_col].value
        if id_val is None:
            continue
        if str(inc_val).strip().lower() in ("yes", "si", "sì", "1", "true"):
            ids.append(int(id_val))
    wb.close()
    return ids


def resolve_arrival_pool(args, client: OdooClient) -> list:
    """
    Return [(employee_id, name)] for the requested pool.
    Raises ValueError if no pool option is provided or if any ID is not active in Odoo.
    """
    active = {emp_id: name for emp_id, name in client.fetch_active_employees()}

    if args.all_employees:
        return list(active.items())

    if args.employees:
        ids = [int(x.strip()) for x in args.employees.split(",") if x.strip()]
    elif args.employees_file:
        ids = _load_pool_from_excel_file(args.employees_file)
    else:
        raise ValueError(
            "Specificare almeno una delle opzioni: --employees, --employees-file, --all-employees"
        )

    missing = [i for i in ids if i not in active]
    if missing:
        raise ValueError(f"ID dipendenti non trovati tra i dipendenti attivi in Odoo: {missing}")

    return [(i, active[i]) for i in ids]


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--mode", choices=["ops", "day", "month", "attendance-day", "arrival-stats"], default="ops",
                        help="ops=oggi+ieri, day=un giorno, month=intero mese, attendance-day=presenze/assenze giorno, arrival-stats=statistiche ingressi")
    parser.add_argument("--lookback-days", type=int, default=1,
                        help="Solo per mode=ops: quanti giorni indietro includere (default 1 = ieri)")
    parser.add_argument("--day", default=None, help="YYYY-MM-DD (richiesto se mode=day, opzionale per attendance-day)")
    parser.add_argument("--also-prev", action="store_true",
                        help="Se mode=day, include anche il giorno precedente")
    parser.add_argument("--month", default=None, help="YYYY-MM (richiesto se mode=month)")
    parser.add_argument("--send-mail", action="store_true", help="Invia email (se mail.enabled=true)")
    parser.add_argument("--no-dedup", action="store_true",
                        help="Non usare deduplica (utile per report mese/giorno)")
    parser.add_argument("--output", default=None,
                        help="Path output Excel (solo mode=attendance-day)")
    # --- arrival-stats ---
    parser.add_argument("--from-day", default=None, metavar="YYYY-MM-DD",
                        help="Inizio periodo (default: 1 gennaio anno corrente)")
    parser.add_argument("--to-day", default=None, metavar="YYYY-MM-DD",
                        help="Fine periodo (default: oggi)")
    parser.add_argument("--arrival-start", default="08:00", metavar="HH:MM",
                        help="Inizio finestra ingresso valido (default: 08:00)")
    parser.add_argument("--arrival-end", default="09:00", metavar="HH:MM",
                        help="Fine finestra ingresso valido (default: 09:00)")
    parser.add_argument("--bucket-minutes", type=int, default=10, metavar="N",
                        help="Ampiezza delle fasce in minuti (default: 10)")
    parser.add_argument("--employees", default=None, metavar="ID,ID,...",
                        help="Pool esplicito: IDs separati da virgola")
    parser.add_argument("--employees-file", default=None, metavar="PATH",
                        help="Pool da file Excel (colonne: employee_id, employee_name, include)")
    parser.add_argument("--all-employees", action="store_true",
                        help="Usa tutti i dipendenti attivi come pool")
    args = parser.parse_args()

    with open("config.yaml", "r", encoding="utf-8") as f:
        cfg = yaml.safe_load(f)

    if args.mode == "attendance-day":
        if cfg["mode"].upper() != "ODOO":
            raise RuntimeError("mode=attendance-day supportato solo con config.yaml mode=ODOO")

        target_day = resolve_target_day(args.day, cfg["timezone"])
        o = cfg["odoo"]
        client = OdooClient(o["url"], o["db"], o["user"], o["password"], cfg["timezone"])

        start_utc, end_utc = day_range_utc(target_day, cfg["timezone"])
        sessions = client.fetch_sessions_for_day(start_utc, end_utc)
        employees = client.fetch_active_employees()

        rows = build_attendance_rows(target_day, employees, sessions)

        output_path = Path(args.output) if args.output else Path("reports") / f"presenze_assenze_{target_day.isoformat()}.xlsx"
        generated = export_attendance_excel(target_day, rows, output_path)
        print(f"Report generato: {generated}")

        mail_cfg = cfg.get("mail", {})
        if args.send_mail:
            if not mail_cfg.get("enabled", False):
                print("Invio email non effettuato: mail.enabled=false in config.yaml")
            else:
                attendance_subject = f"Report presenze e assenze - {target_day}"
                attendance_body = (
                    "Buongiorno,\n\n"
                    f"vi segnaliamo in allegato il report delle presenze e assenze relativo alla giornata del {target_day}.\n\n"
                    "Il file contiene l'elenco dei dipendenti attivi e lo stato della presenza registrata su Odoo.\n\n"
                    "La presente email è stata generata automaticamente dal sistema di controllo presenze.\n\n"
                    "Cordiali saluti."
                )
                send_mail(
                    subject=attendance_subject,
                    body=attendance_body,
                    mail_cfg=mail_cfg,
                    attachments=[str(generated)],
                )
                print("Email inviata.")
        return

    if args.mode == "arrival-stats":
        if cfg["mode"].upper() != "ODOO":
            raise RuntimeError("mode=arrival-stats è supportato solo con config.yaml mode=ODOO")

        try:
            from_day, to_day, arrival_start, arrival_end, bucket_minutes = validate_arrival_args(args)
        except ValueError as exc:
            print(f"Errore parametri: {exc}")
            raise SystemExit(1)

        o = cfg["odoo"]
        client = OdooClient(o["url"], o["db"], o["user"], o["password"], cfg["timezone"])

        try:
            pool = resolve_arrival_pool(args, client)
        except ValueError as exc:
            print(f"Errore pool dipendenti: {exc}")
            raise SystemExit(1)

        print(
            f"Pool: {len(pool)} dipendente/i  |  Periodo: {from_day} - {to_day}"
            f"  |  Finestra: {arrival_start.strftime('%H:%M')}-{arrival_end.strftime('%H:%M')}"
        )

        output_path = Path(args.output) if args.output else default_output_path(from_day, to_day)
        params = ArrivalStatsParams(
            from_day=from_day, to_day=to_day,
            arrival_start=arrival_start, arrival_end=arrival_end,
            bucket_minutes=bucket_minutes,
            employee_pool=pool,
            output_path=output_path,
        )
        result = _svc_run(params, cfg)
        generated = result.output_path
        period_stats = result.period_stats
        print(f"Report generato: {generated}")
        print(f"  Ingressi validi: {period_stats.valid_count}  |  Esclusi: {period_stats.excluded_count}  |  Media: {period_stats.mean_dt}")

        mail_cfg = cfg.get("mail", {})
        if args.send_mail:
            if not mail_cfg.get("enabled", False):
                print("Invio email non effettuato: mail.enabled=false in config.yaml")
            else:
                send_mail(
                    subject=f"Statistiche ingressi {from_day} - {to_day}",
                    body=(
                        f"Report statistiche orari di ingresso allegato.\n\n"
                        f"Periodo: {from_day} - {to_day}\n"
                        f"Pool: {len(pool)} dipendente/i\n"
                        f"Ingressi validi: {period_stats.valid_count}\n"
                        f"Media: {period_stats.mean_dt}\n"
                    ),
                    mail_cfg=mail_cfg,
                    attachments=[str(generated)],
                )
                print("Email inviata.")
        return

    days = sorted(build_days_to_check(args))

    # In ODOO mode: crea client una sola volta
    client = None
    if cfg["mode"].upper() == "ODOO":
        o = cfg["odoo"]
        client = OdooClient(o["url"], o["db"], o["user"], o["password"], cfg["timezone"])

    all_anomalies = []

    for d in days:
        # 🚫 Salta sabato e domenica
        if d.weekday() >= 5:
            print(f"Skip weekend: {d}")
            continue

        # 1) carica sessioni
        if cfg["mode"].upper() == "FAKE":
            sessions = build_fake_sessions(d, cfg["timezone"])
        elif cfg["mode"].upper() == "ODOO":
            start_utc, end_utc = day_range_utc(d, cfg["timezone"])
            sessions = client.fetch_sessions_for_day(start_utc, end_utc)
            employees = client.fetch_active_employees()
        else:
            raise RuntimeError("config.yaml: mode deve essere FAKE o ODOO")

        # 2) detect
        if cfg["mode"].upper() == "ODOO":
            anomalies = detect_for_day(d, sessions, cfg, employees=employees)  # <-- passiamo employees
        else:
            anomalies = detect_for_day(d, sessions, cfg)

        all_anomalies.extend(anomalies)

    # Dedup:
    # - operativo (ops): di default ON
    # - report (day/month): di default OFF
    use_dedup = (args.mode == "ops") and (not args.no_dedup)

    fresh = all_anomalies
    if use_dedup:
        store = DedupStore("anomalies.db")
        store.init()
        tmp = []
        for a in all_anomalies:
            key_day = str(a.day)
            if not store.already_sent(key_day, a.employee_id, a.code):
                store.mark_sent(key_day, a.employee_id, a.code)
                tmp.append(a)
        fresh = tmp

    if not fresh:
        print("Nessuna anomalia trovata.")
        return


    # --- Soppressione NO_CHECKIN se già c'è NO_ATTENDANCE_ALL_DAY (pomeriggio o giorni passati) ---
    tz = ZoneInfo(cfg["timezone"])
    now_local = datetime.now(tz)
    today_local = now_local.date()

    cutoff_str = cfg.get("rules", {}).get("suppress_no_checkin_if_absent_after", "14:30")
    cutoff_h, cutoff_m = map(int, cutoff_str.split(":"))
    after_cutoff = (now_local.hour, now_local.minute) >= (cutoff_h, cutoff_m)

    # chiavi (day, employee_id) per cui esiste assenza tutto il giorno
    absent_all_day_keys = {(a.day, a.employee_id) for a in fresh if a.code == "NO_ATTENDANCE_ALL_DAY"}

    filtered = []
    for a in fresh:
        # Se è un NO_CHECKIN, lo togliamo quando:
        # - lo stesso dipendente/giorno è già marcato come assente tutto il giorno
        # - e (è un giorno passato) oppure (oggi e siamo dopo il cutoff)
        if a.code == "NO_CHECKIN_BY_0900" and (a.day, a.employee_id) in absent_all_day_keys:
            if (a.day < today_local) or (a.day == today_local and after_cutoff):
                continue
        filtered.append(a)

    fresh = filtered
    # --- fine soppressione ---

    # ordina output
    fresh.sort(key=lambda a: (a.day, a.employee_name, a.code))

    body = format_anomalies(fresh)
    print(body)

    # email opzionale
    mail_cfg = cfg.get("mail", {})
    if args.send_mail and mail_cfg.get("enabled", False):
        send_mail(
            subject=f"Odoo – {len(fresh)} anomalie ({days[0]} → {days[-1]})",
            body=body,
            mail_cfg=mail_cfg
        )
        print("\nEmail inviata.")


if __name__ == "__main__":
    main()
